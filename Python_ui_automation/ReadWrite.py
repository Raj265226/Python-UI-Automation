import csv
import json
from openpyxl import load_workbook
import pytest
from selenium import webdriver
from selenium.webdriver.common.by import By

@pytest.fixture(scope="module")
def driver():
    driver = webdriver.Edge()
    driver.maximize_window()
    yield driver
    driver.quit()

def test_csv(driver):
    rows = []

    with open("ReadWriteFile/users.csv", "r") as file:
        reader = csv.DictReader(file)

        for row in reader:
            driver.get("https://demoqa.com/text-box")
            username = row["Username"]
            email = row["Email"]
            driver.find_element(By.ID, "userName").clear()
            driver.find_element(By.ID, "userName").send_keys(username)
            driver.find_element(By.ID, "userEmail").clear()
            driver.find_element(By.ID, "userEmail").send_keys(email)
            actual_username = driver.find_element(By.ID, "userName").get_attribute("value")
            actual_email = driver.find_element(By.ID, "userEmail").get_attribute("value")
            row["Status"] = ("Pass" if actual_username == username and actual_email == email else "Fail")
            rows.append(row)

    with open("ReadWriteFile/users.csv", "w", newline="") as file:
        writer = csv.DictWriter(file,fieldnames=["Username", "Email", "Status"])
        writer.writeheader()
        writer.writerows(rows)

def test_json(driver):
    with open("ReadWriteFile/users.json", "r") as file:
        users = json.load(file)

    for user in users:
        driver.get("https://demoqa.com/text-box")
        driver.find_element(By.ID, "userName").clear()
        driver.find_element(By.ID, "userName").send_keys(user["Username"])
        driver.find_element(By.ID, "userEmail").clear()
        driver.find_element(By.ID, "userEmail").send_keys(user["Email"])
        actual_username = driver.find_element(By.ID, "userName").get_attribute("value")
        actual_email = driver.find_element(By.ID, "userEmail").get_attribute("value")
        user["Status"] = ("Pass" if actual_username == user["Username"] and actual_email == user["Email"] else "Fail")

    with open("ReadWriteFile/users.json", "w") as file:
        json.dump(users, file, indent=4)

def test_excel(driver):
    workbook = load_workbook("ReadWriteFile/users.xlsx")
    sheet = workbook.active
    headers = {}

    for col in range(1, sheet.max_column + 1):
        headers[sheet.cell(1, col).value] = col

    username_col = headers["Username"]
    email_col = headers["Email"]
    status_col = headers["Status"]

    for row in range(2, sheet.max_row + 1):
        username = sheet.cell(row, username_col).value
        email = sheet.cell(row, email_col).value
        driver.get("https://demoqa.com/text-box")
        driver.find_element(By.ID, "userName").clear()
        driver.find_element(By.ID, "userName").send_keys(username)
        driver.find_element(By.ID, "userEmail").clear()
        driver.find_element(By.ID, "userEmail").send_keys(email)
        actual_username = driver.find_element(By.ID, "userName").get_attribute("value")
        actual_email = driver.find_element(By.ID, "userEmail").get_attribute("value")

        sheet.cell(row, status_col).value = ("Pass" if actual_username == username and actual_email == email else "Fail")

    workbook.save("ReadWriteFile/users.xlsx")


def get_test_data_by_setup(file_path, setup_name):
    rows = []
    selected_row = None

    with open(file_path, "r", newline="") as file:
        reader = csv.DictReader(file)
        fieldnames = reader.fieldnames

        for row in reader:
            if row["Name"] == setup_name:
                selected_row = row
            rows.append(row)

    if selected_row is None:
        raise ValueError(f"Setup name '{setup_name}' not found in CSV")
    return rows, fieldnames, selected_row


def update_status_in_csv(file_path, rows, fieldnames, setup_name, status):
    for row in rows:
        if row["Name"] == setup_name:
            row["Status"] = status
            break

    with open(file_path, "w", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def test_csv_selected_setup(driver):
    file_path = "ReadWriteFile/TDD_users.csv"
    setup_name = "non-active"

    rows, fieldnames, current_row = get_test_data_by_setup(file_path, setup_name)
    driver.get("https://demoqa.com/text-box")
    username = current_row["Username"]
    email = current_row["Email"]
    driver.find_element(By.ID, "userName").send_keys(username)
    driver.find_element(By.ID, "userEmail").send_keys(email)
    actual_username = driver.find_element(By.ID, "userName").get_attribute("value")
    actual_email = driver.find_element(By.ID, "userEmail").get_attribute("value")

    if actual_username == username and actual_email == email:
        status = "Pass"
    else:
        status = "Fail"

    update_status_in_csv(file_path,rows,fieldnames,setup_name,status)

CSV_FILE = "ReadWriteFile/DP_users.csv"
def read_csv_data():
    test_data = []

    with open(CSV_FILE, "r", newline="") as file:
        reader = csv.DictReader(file)

        for row in reader:
            test_data.append(row)
    return test_data


def update_csv_status(username, email, status):
    rows = []

    with open(CSV_FILE, "r", newline="") as file:
        reader = csv.DictReader(file)
        fieldnames = reader.fieldnames

        for row in reader:
            if row["Username"] == username and row["Email"] == email:
                row["Status"] = status
            rows.append(row)

    with open(CSV_FILE, "w", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

@pytest.mark.parametrize("current_row", read_csv_data())
def test_csv_dataprovider(driver, current_row):
    driver.get("https://demoqa.com/text-box")
    username = current_row["Username"]
    email = current_row["Email"]
    driver.find_element(By.ID, "userName").send_keys(username)
    driver.find_element(By.ID, "userEmail").send_keys(email)
    actual_username = driver.find_element(By.ID, "userName").get_attribute("value")
    actual_email = driver.find_element(By.ID, "userEmail").get_attribute("value")

    if actual_username == username and actual_email == email:
        status = "Pass"
    else:
        status = "Fail"

    update_csv_status(username, email, status)
    assert status == "Pass", (f"Validation failed for Username: {username}, Email: {email}")